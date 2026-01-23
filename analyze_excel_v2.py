import pandas as pd
import openpyxl

file_path = 'emmvee sample data.xlsx'

try:
    print("--- PANDAS ANALYSIS ---")
    # Read ignoring header first to see raw structure if needed, but defaults usually work
    df = pd.read_excel(file_path)
    print(f"Shape: {df.shape}")
    print("Columns:", list(df.columns)[:20]) # First 20 cols
    
    print("\nLast 5 Rows (First 15 Columns):")
    print(df.iloc[-5:, :15].to_string())

    print("\n--- OPENPYXL FORMULA CHECK ---")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active
    
    # Check the last 3 rows for formulas
    max_row = ws.max_row
    max_col = min(ws.max_column, 20) # Check first 20 cols
    
    start_row = max(1, max_row - 2)
    
    for r in range(start_row, max_row + 1):
        row_values = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                 row_values.append(f"FORMULA:{val}")
            else:
                 row_values.append(str(val))
        print(f"Row {r}: {row_values}")

except Exception as e:
    print(f"Error: {e}")
