import pandas as pd
import openpyxl

file_path = 'emmvee sample data.xlsx'

try:
    # Read using openpyxl to get formulas if possible, but pandas is easier for data
    # Let's inspect the last few rows using pandas first
    df = pd.read_excel(file_path)
    print("Columns:", df.columns.tolist())
    print("\nLast 5 rows:")
    print(df.tail(5))
    
    # Check for a 'Total' row by looking for 'Total' string in any column
    # Often it's in the first column or 'Shift' column
    print("\nSearching for 'Total' row...")
    # Convert all to string and search
    mask = df.apply(lambda x: x.astype(str).str.contains('Total', case=False).any(), axis=1)
    total_rows = df[mask]
    if not total_rows.empty:
        print("Found Total Rows:")
        print(total_rows)
    else:
        print("No explicit 'Total' string found in data rows.")
        
    # Also load with openpyxl to see if there are formulas
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active
    print("\nInspecting formulas in the last row (if any):")
    max_row = ws.max_row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=max_row, column=col)
        # Check if it has a formula
        if isinstance(cell.value, str) and cell.value.startswith('='):
             print(f"Column {col} ({cell.column_letter}): Formula = {cell.value}")
        else:
             print(f"Column {col} ({cell.column_letter}): Value = {cell.value}")

except Exception as e:
    print(f"Error: {e}")
