import openpyxl
import sys

file_path = 'emmvee sample data.xlsx'

try:
    # Set stdout to utf-8 just in case
    sys.stdout.reconfigure(encoding='utf-8')
    
    wb = openpyxl.load_workbook(file_path, data_only=False)
    # Print all sheet names
    print(f"Sheets: {wb.sheetnames}")
    
    ws = wb.active # or wb['Sheet1'] if known
    print(f"Active Sheet: {ws.title}")
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    print(f"Max Row: {max_row}, Max Col: {max_col}")
    
    # Inspect the last 5 rows for any formulas or "Total" labels
    start_row = max(1, max_row - 5)
    
    for r in range(start_row, max_row + 2): # Go a bit beyond just in case
        row_vals = []
        has_formula = False
        row_has_total = False
        
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            
            if isinstance(val, str):
                if val.startswith('='):
                    row_vals.append(f"FORMULA: {val}")
                    has_formula = True
                elif 'total' in val.lower():
                    row_vals.append(f"LABEL: {val}")
                    row_has_total = True
                else:
                    row_vals.append(str(val))
            else:
                 row_vals.append(str(val))
        
        if has_formula or row_has_total:
            print(f"Row {r}: {row_vals}")

except Exception as e:
    print(f"Error: {e}")
