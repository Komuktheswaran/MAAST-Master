import openpyxl
import sys

file_path = 'emmvee sample data.xlsx'

try:
    sys.stdout.reconfigure(encoding='utf-8')
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active
    
    max_row = ws.max_row
    # Force max col to 20 for readability
    max_col = 20
    
    start_row = max(1, max_row - 7) # Look at last 7 rows
    
    print(f"Reading rows {start_row} to {max_row} (First 20 columns)")
    
    for r in range(start_row, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                 row_vals.append(f"FORMULA:{val}")
            else:
                 row_vals.append(str(val))
        print(f"Row {r}: {row_vals}")

except Exception as e:
    print(f"Error: {e}")
